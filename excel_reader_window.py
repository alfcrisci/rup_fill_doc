import os
import sys
from datetime import datetime
from docxtpl import DocxTemplate
import openpyxl
from openpyxl.utils import get_column_letter

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QFileDialog, QTextEdit, QMessageBox,
    QListWidget, QListWidgetItem, QGroupBox, QTabWidget,
    QFormLayout, QSplitter, QFrame, QScrollArea
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QPixmap, QIcon
from PyQt6.QtWidgets import QApplication

class ExcelReaderWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("RUP IBE helper - Document Generator")
        self.setGeometry(100, 100, 1300, 800)
        
        # Set application icon (replace with your icon if needed)
        self.setWindowIcon(QIcon(":/images/app_icon.png"))
        # Variabili per le cartelle predefinite
        self.cartella_assoluta = os.path.expanduser("~")  # Cartella home utente come default
        self.cartella_template = os.path.expanduser("~")  # Cartella home utente come default
      
        # Central widget with tabs
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # =============================================
        # HEADER SECTION (Customizable Title, Logo, Credits)
        # =============================================
        header_frame = QFrame()
        header_frame.setFrameShape(QFrame.Shape.StyledPanel)
        header_frame.setStyleSheet("background-color: #f8f9fa; border-bottom: 1px solid #dee2e6;")
        header_layout = QHBoxLayout()
        header_frame.setLayout(header_layout)
        
        # Left section for logo
        self.logo_label = QLabel()
        self.logo_label.setFixedSize(320, 120)
        self.logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.logo_label.setStyleSheet("border: none;")
        header_layout.addWidget(self.logo_label)
        
        # Center section for title
        title_frame = QFrame()
        title_layout = QVBoxLayout()
        title_frame.setLayout(title_layout)
        
        self.title_label = QLabel("RUP IBE HELPER")
        self.title_label.setStyleSheet("""
            font-size: 28px; 
            font-weight: bold; 
            color: #2c3e50;
            margin-bottom: 5px;
        """)
        self.title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.subtitle_label = QLabel("Generatore di documenti per procedure d'acquisto")
        self.subtitle_label.setStyleSheet("""
            font-size: 14px; 
            color: #7f8c8d;
            font-style: italic;
        """)
        self.subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        title_layout.addWidget(self.title_label)
        title_layout.addWidget(self.subtitle_label)
        header_layout.addWidget(title_frame, stretch=1)
        
        # Right section for credits
        credits_frame = QFrame()
        credits_layout = QVBoxLayout()
        credits_frame.setLayout(credits_layout)
        
        self.version_label = QLabel("Versione 1.0")
        self.author_label = QLabel("Sviluppato da: IBE CNR")
        self.date_label = QLabel(datetime.now().strftime("%d/%m/%Y"))
        
        for label in [self.version_label, self.author_label, self.date_label]:
            label.setStyleSheet("""
                font-size: 12px; 
                color: #7f8c8d;
                margin: 2px;
            """)
            label.setAlignment(Qt.AlignmentFlag.AlignRight)
            credits_layout.addWidget(label)
        
        header_layout.addWidget(credits_frame)
        
        # Add header to main layout
        
        main_layout.addWidget(header_frame)
        
        # Separator line
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)
        separator.setStyleSheet("margin: 5px 0; border-color: #dee2e6;")
        main_layout.addWidget(separator)
        
        # =============================================
        # MAIN CONTENT (Tabs)
        # =============================================
        # Create tabs
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabBar::tab {
                padding: 8px;
                min-width: 120px;
            }
            QTabBar::tab:selected {
                background: #e9ecef;
                border-bottom: 2px solid #2c3e50;
            }
        """)
        main_layout.addWidget(self.tabs)
        
        # Excel Reader Tab
        self.setup_excel_tab()
        
        # Unified Document Generator Tab
        self.setup_unified_document_tab()
        
        # Current data storage
        self.current_file = None
        self.sheet_data = {
            'dati_generali_procedura': None,
            'generazioni_offerte': None
        }
        
        # Load default logo (replace with your logo path)
        self.load_logo("images/default_logo.png")

    def load_logo(self, path):
        """Load logo image with fallback to text"""
        try:
            pixmap = QPixmap(path)
            if not pixmap.isNull():
                self.logo_label.setPixmap(pixmap.scaled(
                    self.logo_label.size(),
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation
                ))
                return
        except:
            pass
        
        # Fallback text logo
        self.logo_label.setText("LOGO")
        self.logo_label.setStyleSheet("""
            background-color: #e9ecef;
            border-radius: 4px;
            font-weight: bold;
            font-size: 24px;
            color: #2c3e50;
        """)

    def setup_excel_tab(self):
        """Setup the Excel reader tab"""
        excel_tab = QWidget()
        excel_layout = QVBoxLayout()
        excel_layout.setContentsMargins(10, 10, 10, 10)
        excel_layout.setSpacing(10)
        excel_tab.setLayout(excel_layout)
        
        # File selection
        file_group = QGroupBox("Selezione File Excel")
        file_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                padding: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
        """)
        file_layout = QHBoxLayout()
        file_layout.setContentsMargins(5, 5, 5, 5)
        self.file_path = QLineEdit()
        self.file_path.setPlaceholderText("Seleziona file Excel...")
        file_button = QPushButton("Sfoglia...")
        file_button.setIcon(QIcon.fromTheme("document-open"))
        file_button.clicked.connect(self.browse_excel_file)
        file_layout.addWidget(self.file_path)
        file_layout.addWidget(file_button)
        file_group.setLayout(file_layout)
        
        # Splitter for the two lists
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.splitter.setHandleWidth(8)
        self.splitter.setStyleSheet("""
            QSplitter::handle {
                background: #dee2e6;
                width: 4px;
                margin: 2px;
            }
        """)
        
        # List for dati_generali_procedura
        self.dati_generali_list = QListWidget()
        self.dati_generali_list.itemDoubleClicked.connect(
            lambda item: self.show_variable_value(item, 'dati_generali_procedura'))
        self.dati_generali_list.setMinimumWidth(300)
        
        # List for generazioni_offerte
        self.generazioni_offerte_list = QListWidget()
        self.generazioni_offerte_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        self.generazioni_offerte_list.itemDoubleClicked.connect(
            lambda item: self.show_variable_value(item, 'generazioni_offerte'))
        self.generazioni_offerte_list.setMinimumWidth(300)
        
        # Add lists to splitter
        self.splitter.addWidget(self.create_list_group(self.dati_generali_list, "Dati Procedura"))
        self.splitter.addWidget(self.create_list_group(self.generazioni_offerte_list, "Generazioni Richeste Offerte - Selezione multipla con click"))
        
        # Set initial sizes and stretch factors
        self.splitter.setSizes([400, 400])
        self.splitter.setStretchFactor(0, 1)
        self.splitter.setStretchFactor(1, 1)
        
        # Results display
        self.results_display = QTextEdit()
        self.results_display.setReadOnly(True)
        self.results_display.setStyleSheet("""
            QTextEdit {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                padding: 8px;
                min-height: 120px;
                font-size: 12px;
            }
        """)
        
        # Scan button
        scan_button = QPushButton("Leggi Fogli Excel")
        scan_button.setIcon(QIcon.fromTheme("document-open"))
        scan_button.setStyleSheet("""
            QPushButton {
                padding: 8px;
                background-color: #2c3e50;
                color: white;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #34495e;
            }
        """)
        scan_button.clicked.connect(self.read_excel_sheets)
        
        # Add widgets to Excel tab
        excel_layout.addWidget(file_group)
        excel_layout.addWidget(scan_button)
        excel_layout.addWidget(self.splitter)
        excel_layout.addWidget(QLabel("Output:"))
        excel_layout.addWidget(self.results_display)
        
        self.tabs.addTab(excel_tab, "Excel Reader")

    def create_list_group(self, list_widget, title):
        """Create a group box for a list widget"""
        group = QGroupBox(title)
        group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                margin: 5px;
                padding: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
        """)
        layout = QVBoxLayout()
        layout.setContentsMargins(5, 5, 5, 5)
        
        list_widget.setStyleSheet("""
            QListWidget {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                padding: 5px;
                font-size: 12px;
            }
            QListWidget::item {
                padding: 3px;
            }
            QListWidget::item:selected {
                background-color: #d4e6f1;
                color: #2c3e50;
            }
        """)
        
        layout.addWidget(list_widget)
        group.setLayout(layout)
        return group

    def browse_excel_file(self):
        """Open a file dialog to select an Excel file"""
        file_name, _ = QFileDialog.getOpenFileName(
            self, 
            "Apri file Excel", 
            self.cartella_assoluta,  # Usa la cartella predefinita
            "Excel Files (*.xlsx *.xls)"
        )
        if file_name:
            self.file_path.setText(file_name)
            self.current_file = file_name
            self.cartella_assoluta = os.path.dirname(file_name)  # Aggiorna la cartella con l'ultima usata
            self.results_display.append(f"File selezionato: {file_name}")

    def format_excel_date(self, value):
        """Format Excel date value to dd/mm/YYYY string"""
        if value is None:
            return ""
        
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        elif isinstance(value, str):
            try:
                # Try to parse string as date
                dt = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                return dt.strftime("%d/%m/%Y")
            except ValueError:
                try:
                    dt = datetime.strptime(value, "%d/%m/%Y")
                    return dt.strftime("%d/%m/%Y")
                except ValueError:
                    return value
        else:
            try:
                # Try to convert Excel numeric date
                dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(value) - 2)
                return datetime.fromordinal(dt).strftime("%d/%m/%Y")
            except (ValueError, TypeError):
                return str(value)

    def read_excel_sheets(self):
        if not self.current_file:
            QMessageBox.warning(self, "Attenzione", "Seleziona prima un file excel.")
            return
            
        try:
            workbook = openpyxl.load_workbook(self.current_file, read_only=True, data_only=True)
            
            # Clear previous data
            self.dati_generali_list.clear()
            self.generazioni_offerte_list.clear()
            self.sheet_data = {
                'dati_generali_procedura': None,
                'generazioni_offerte': None
            }
            
            # List of date fields that need special formatting
            date_fields = [
                'data_nascita_richiedente',
                'data_nascita_RUP',
                'data_nascita_direttore',
                'data_nascita_RSS',
                'data_rda',
                'data_scadenza',
                'data_scadenza_offerta',
                'data_oggi'
            ]
            
            # Read dati_generali_procedura sheet
            if 'dati_generali_procedura' in workbook.sheetnames:
                sheet = workbook['dati_generali_procedura']
                self.sheet_data['dati_generali_procedura'] = sheet
                
                # Mappatura tra nomi delle colonne Excel e campi QLineEdit
                field_mapping = {
                    'numero_CUP': None,
                    'servizio_fornitura': None,
                    'acronimo_progetto': None,
                    'oggetto_fornitura_servizio': None,
                    'oggetto_esteso_fornitura_servizio': None,
                    'nome_cognome_richiedente': None,
                    'mail_contatto_richiedente': None,
                    'data_oggi': None,
                    'data_rda': None,
                    'data_scadenza': None,
                    'descrizione_servizio_fornitura': None,
                    'breve_descrizione_caratteristiche_prestazioni_acquisizione_bene_servizio': None,
                    'breve_descrizione_motivazione_acquisizione_bene_servizio': None,
                    'clausola_cam': None,
                    'clausola_servizi_fornitura': None,
                    'dichiarazione_deroga_MEPA': None,
                    'dichiarazione_mancada_consip_informatica': None,
                    'dichiarazione_valore_affidamento': None,
                    'dichiarazione_motivo_deroga_principio_rotazione': None,
                    'importo_massimo': None,
                    'quantita': None,
                    'numero_CIG': None,
                    'numero_COAN': None,
                    'voce_piano_dei_conti': None,
                    'voce_costo_COAN': None,
                    'codice_CPV': None,
                    'piattaforma_scelta': None,
                    'bando_MEPA':None,
                    'riferimento_PAD': None,
                    'codice_ateco_OE': None,
                    'codice_ateco_OE_sec': None,
                    'codice_ateco_OE_dich': None,
                    'dichiarazione_individuazione_OE': None,
                    'indirizzo_OE_scelta': None,
                    'legale_rap_OE_scelta': None,
                    'sede_OE_scelta': None,
                    'piva_OE_scelta': None,
                    'codice_CNEL': None,
                    'estratti_CNEL': None,
                    'data_nascita_richiedente': None,
                    'luogo_nascita_richiedente': None,
                    'CF_richiedente': None,
                    'sede_richiedente': None,
                    'dichiarazioni_comunicazione_incarichi_richiedente': None,
                    'dichiarazioni_partecipazione_associazioni_organizzazioni_richiedente': None,
                    'qualifica_richiedente': None,
                    'nome_cognome_RUP': None,
                    'data_nascita_RUP': None,
                    'luogo_nascita_RUP': None,
                    'CF_RUP': None,
                    'sede_RUP': None,
                    'mail_contatto_RUP': None,
                    'nome_cognome_direttore': None,
                    'data_nascita_direttore': None,
                    'luogo_nascita_direttore': None,
                    'CF_direttore': None,
                    'sede_direttore': None,
                    'mail_contatto_direttore': None,
                    'nome_cognome_RSS': None,
                    'data_nascita_RSS': None,
                    'luogo_nascita_RSS': None,
                    'CF_RSS': None,
                    'sede_RSS': None,
                    'mail_contatto_RSS': None,
                    'ulteriori_riferimenti_normativi_attuativi_operativi': None,
                    'url_gara': None,
                    'protocollo_RDA': None,
                    'protocollo_richiesta_url': None,
                    'protocollo_nomina_RUP': None,
                    'protocollo_conflittoint_richiedente': None,
                    'protocollo_conflittoint_RUP': None,
                    'protocollo_conflittoint_direttore': None,
                    'protocollo_allegato2_CIG': None,
                    'protocollo_istruttoria_RUP': None,
                    'protocollo_DAC': None,
                    'protocollo_ordine': None
                }
                
                # Process column C with names from column E, starting from row 2
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=3, max_col=5, values_only=False), start=2):
                    cell_c = row[0]  # Column C
                    cell_e = row[2]  # Column E
                    
                    if (cell_c.value is None or not str(cell_c.value).strip() or 
                        cell_e.value is None or not str(cell_e.value).strip()):
                        continue
                        
                    name = str(cell_e.value)
                    value = cell_c.value
                    
                    # Format date fields
                    for date_field in date_fields:
                        if date_field.lower() in name.lower():
                            value = self.format_excel_date(value)
                            break
                    
                    item_text = f"{name}: {value}"
                    
                    # Check if this value matches one of our fields
                    for field_name in field_mapping.keys():
                        if field_name.lower() in name.lower():
                            field_mapping[field_name] = str(value) if value else ""
                    
                    item = QListWidgetItem(item_text)
                    item.setData(Qt.ItemDataRole.UserRole, {
                        'type': 'value',
                        'coord': cell_c.coordinate,
                        'name': name,
                        'value': value
                    })
                    self.dati_generali_list.addItem(item)
                
                # Process column D with flag names, starting from row 2
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=False), start=2):
                    cell_d = row[0]  # Column D
                    
                    if cell_d.value is not None and str(cell_d.value).strip():
                        flag_name = f"flag_{row_idx}"
                        item_text = f"{flag_name}: {cell_d.value}"
                        
                        item = QListWidgetItem(item_text)
                        item.setData(Qt.ItemDataRole.UserRole, {
                            'type': 'flag',
                            'coord': cell_d.coordinate,
                            'name': flag_name,
                            'value': cell_d.value
                        })
                        self.dati_generali_list.addItem(item)
                
                # Auto-fill QLineEdit fields
                for field_name, value in field_mapping.items():
                    if value is not None and field_name in self.doc_fields:
                        self.doc_fields[field_name].setText(value)
            
            # Read generazioni_offerte sheet as table (first row = column names)
            if 'generazioni_offerte' in workbook.sheetnames:
                sheet = workbook['generazioni_offerte']
                self.sheet_data['generazioni_offerte'] = sheet
                
                # Read first row as column names
                headers = []
                for cell in sheet[1]:
                    header_name = str(cell.value) if cell.value else f"col_{cell.column_letter}"
                    headers.append(header_name)
                
                # Read data for each subsequent row
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    row_data = {}
                    for col_idx, cell in enumerate(row, start=1):
                        if col_idx - 1 < len(headers):
                            value = cell.value
                            # Format date fields
                            for date_field in date_fields:
                                if date_field.lower() in headers[col_idx - 1].lower():
                                    value = self.format_excel_date(value)
                                    break
                            row_data[headers[col_idx - 1]] = value
                    
                    # Add to list as single item for the row
                    item_text = f" {', '.join(f' {v}' for k, v in row_data.items() if v)}"
                    item = QListWidgetItem(item_text)
                    item.setData(Qt.ItemDataRole.UserRole, {
                        'type': 'row',
                        'row_idx': row_idx-1,
                        'data': row_data
                    })
                    self.generazioni_offerte_list.addItem(item)
            
            workbook.close()
            
        except Exception as e:
            QMessageBox.critical(self, "Errore", f"Errore nella lettura del file Excel:\n{str(e)}")

    def show_variable_value(self, item, sheet_name):
        if not self.current_file or sheet_name not in self.sheet_data or not self.sheet_data[sheet_name]:
            return
            
        item_data = item.data(Qt.ItemDataRole.UserRole)
        
        try:
            if sheet_name == 'generazioni_offerte' and item_data['type'] == 'row':
                # Show all row data
                row_data = item_data['data']
                details = "\n".join([f"• {k}: {v} ({type(v).__name__})" for k, v in row_data.items()])
                self.results_display.append(
                    f"\nDettaglio riga {item_data['row_idx']}:\n"
                    f"{details}\n"
                    f"----------------------------")
            else:
                # Show details for dati_generali_procedura or single cells
                sheet = self.sheet_data[sheet_name]
                
                if item_data['type'] == 'value':
                    self.results_display.append(
                        f"\nDettaglio variabile (valore):\n"
                        f"• Nome: {item_data['name']}\n"
                        f"• Posizione: {item_data['coord']}\n"
                        f"• Valore: {item_data['value']}\n"
                        f"• Tipo: {type(item_data['value']).__name__}\n"
                        f"----------------------------")
                
                elif item_data['type'] == 'flag':
                    self.results_display.append(
                        f"\nDettaglio flag:\n"
                        f"• Nome: {item_data['name']}\n"
                        f"• Posizione: {item_data['coord']}\n"
                        f"• Valore: {item_data['value']}\n"
                        f"• Tipo: {type(item_data['value']).__name__}\n"
                        f"----------------------------")
                
        except Exception as e:
            QMessageBox.critical(self, "Errore", f"Errore nella lettura della cella:\n{str(e)}")

    def setup_unified_document_tab(self):
        """Setup a unified Document Generator tab with scrollable content"""
        doc_tab = QWidget()
        doc_layout = QVBoxLayout()
        doc_layout.setContentsMargins(10, 10, 10, 10)
        doc_layout.setSpacing(10)
        doc_tab.setLayout(doc_layout)
        
        # Create a scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        
        # Create a container widget for the scroll area
        container = QWidget()
        container_layout = QVBoxLayout()
        container_layout.setContentsMargins(5, 5, 5, 5)
        container.setLayout(container_layout)
        
        # Template selection
        template_group = QGroupBox("Configurazione Documento")
        template_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        template_layout = QFormLayout()
        
        self.template_path = QLineEdit()
        self.template_path.setPlaceholderText("Seleziona template Word (multipli)...")
        template_btn = QPushButton("Sfoglia...")
        template_btn.setIcon(QIcon.fromTheme("document-open"))
        template_btn.clicked.connect(self.browse_template_file)
        
        self.output_dir = QLineEdit()
        self.output_dir.setPlaceholderText("Cartella di output...")
        output_btn = QPushButton("Sfoglia...")
        output_btn.setIcon(QIcon.fromTheme("folder-open"))
        output_btn.clicked.connect(self.browse_output_dir)
        
        template_layout.addRow("Template Word:", self.template_path)
        template_layout.addRow(template_btn)
        template_layout.addRow("Cartella Output:", self.output_dir)
        template_layout.addRow(output_btn)
        
        template_group.setLayout(template_layout)
        container_layout.addWidget(template_group)
        
        # Excel file section
        # excel_group = QGroupBox("File Excel")
        # excel_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        # excel_layout = QHBoxLayout()
        
        # self.excel_path = QLineEdit()
        # self.excel_path.setPlaceholderText("Seleziona file Excel...")
        # excel_btn = QPushButton("Sfoglia...")
        # excel_btn.setIcon(QIcon.fromTheme("document-open"))
        # excel_btn.clicked.connect(self.browse_excel_file)
        
        # excel_layout.addWidget(self.excel_path)
        # excel_layout.addWidget(excel_btn)
        # excel_group.setLayout(excel_layout)
        
        # container_layout.addWidget(excel_group)
        
        # Generate buttons
        buttons_frame = QFrame()
        buttons_layout = QHBoxLayout()
        buttons_layout.setContentsMargins(0, 10, 0, 10)
        buttons_frame.setLayout(buttons_layout)
        
        # Button for dati_generali_procedura
        generate_dati_btn = QPushButton("Genera Documenti Dati Generali")
        generate_dati_btn.setIcon(QIcon.fromTheme("document-save-as"))
        generate_dati_btn.setStyleSheet("""
            QPushButton {
                padding: 8px;
                background-color: #2980b9;
                color: white;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: #3498db;
            }
        """)
        generate_dati_btn.clicked.connect(lambda: self.generate_document('dati_generali_procedura'))
        
        # Button for generazioni_offerte
        generate_offerte_btn = QPushButton("Genera Documenti Offerte")
        generate_offerte_btn.setIcon(QIcon.fromTheme("document-save-as"))
        generate_offerte_btn.setStyleSheet("""
            QPushButton {
                padding: 8px;
                background-color: #27ae60;
                color: white;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
        """)
        generate_offerte_btn.clicked.connect(lambda: self.generate_document('generazioni_offerte'))
        
        buttons_layout.addWidget(generate_dati_btn)
        buttons_layout.addWidget(generate_offerte_btn)
        container_layout.addWidget(buttons_frame)
        
        # Document fields
        self.doc_fields = {
            # Date
            'data_oggi': QLineEdit(),
            'data_rda': QLineEdit(),
            'data_scadenza': QLineEdit(),
            'data_scadenza_offerta': QLineEdit(),
            
            # Informazioni generali
            
            'servizio_fornitura': QLineEdit(),
            'descrizione_servizio_fornitura': QLineEdit(),
            'breve_descrizione_caratteristiche_prestazioni_acquisizione_bene_servizio': QLineEdit(),
            'breve_descrizione_motivazione_acquisizione_bene_servizio': QLineEdit(),
            'oggetto_fornitura_servizio': QLineEdit(),
            'oggetto_esteso_fornitura_servizio': QLineEdit(),
            
            # Valori economici
            'importo_massimo': QLineEdit(),
            'quantita': QLineEdit(),
            'importo_oneri_sicurezza': QLineEdit(),
            'importo_oneri_personale': QLineEdit(),
            
            # Codici e numeri
            'acronimo_progetto': QLineEdit(),
            'numero_CUP': QLineEdit(),
            'numero_CIG': QLineEdit(),
            'numero_COAN': QLineEdit(),
            'voce_piano_dei_conti': QLineEdit(),
            'voce_costo_COAN': QLineEdit(),
            'codice_CPV': QLineEdit(),
            'codice_ateco_OE': QLineEdit(),
            'codice_ateco_OE_sec': QLineEdit(),
            'codice_ateco_OE_dich': QLineEdit(),
            'codice_CNEL': QLineEdit(),
            'estratti_CNEL': QLineEdit(),
            
            # Informazioni OE
            'piattaforma_scelta': QLineEdit(),
            'bando_MEPA': QLineEdit(),
            'riferimento_PAD': QLineEdit(),
            'dichiarazione_individuazione_OE': QLineEdit(),
            'indirizzo_OE_scelta': QLineEdit(),
            'legale_rap_OE_scelta': QLineEdit(),
            'sede_OE_scelta': QLineEdit(),
            'piva_OE_scelta': QLineEdit(),
            
            # Clausole e dichiarazioni
            'clausola_cam': QLineEdit(),
            'clausola_servizi_fornitura': QLineEdit(),
            'dichiarazione_deroga_MEPA': QLineEdit(),
            'dichiarazione_mancata_consip_informatica': QLineEdit(),
            'dichiarazione_valore_affidamento': QLineEdit(),
            'dichiarazione_motivo_deroga_principio_rotazione': QLineEdit(),
       
            # Richiedente
            'nome_cognome_richiedente': QLineEdit(),
            'data_nascita_richiedente': QLineEdit(),
            'luogo_nascita_richiedente': QLineEdit(),
            'CF_richiedente': QLineEdit(),
            'sede_richiedente': QLineEdit(),
            'firma_richiedente': QLineEdit(),
            'img_documento_richiedente': QLineEdit(),
            'dichiarazioni_comunicazione_incarichi_richiedente': QLineEdit(),
            'dichiarazioni_partecipazione_associazioni_organizzazioni_richiedente': QLineEdit(),
            'mail_contatto_richiedente': QLineEdit(),
            'qualifica_richiedente': QLineEdit(),
            
            # RUP
            'nome_cognome_RUP': QLineEdit(),
            'data_nascita_RUP': QLineEdit(),
            'luogo_nascita_RUP': QLineEdit(),
            'CF_RUP': QLineEdit(),
            'sede_RUP': QLineEdit(),
            'firma_RUP': QLineEdit(),
            'img_documento_RUP': QLineEdit(),
            'dichiarazioni_comunicazione_incarichi_RUP': QLineEdit(),
            'dichiarazioni_partecipazione_associazioni_organizzazioni_RUP': QLineEdit(),
            'mail_contatto_RUP': QLineEdit(),
            # RUP
            'nome_cognome_supportoRUP': QLineEdit(),
            'data_nascita_supportoRUP': QLineEdit(),
            'luogo_nascita_supportoRUP': QLineEdit(),
            'CF_supportoRUP': QLineEdit(),
            'sede_supportoRUP': QLineEdit(),
            'firma_supportoRUP': QLineEdit(),
            'img_documento_supportoRUP': QLineEdit(),
            'dichiarazioni_comunicazione_incarichi_supportoRUP': QLineEdit(),
            'dichiarazioni_partecipazione_associazioni_organizzazioni_supportoRUP': QLineEdit(),
            'mail_contatto_supportoRUP': QLineEdit(),
            
            # Direttore
            'nome_cognome_direttore': QLineEdit(),
            'data_nascita_direttore': QLineEdit(),
            'luogo_nascita_direttore': QLineEdit(),
            'CF_direttore': QLineEdit(),
            'sede_direttore': QLineEdit(),
            'firma_direttore': QLineEdit(),
            'img_documento_direttore': QLineEdit(),
            'dichiarazioni_comunicazione_incarichi_direttore': QLineEdit(),
            'dichiarazioni_partecipazione_associazioni_organizzazioni_direttore': QLineEdit(),
            'mail_contatto_direttore': QLineEdit(),
            
            # RSS
            'nome_cognome_RSS': QLineEdit(),
            'data_nascita_RSS': QLineEdit(),
            'luogo_nascita_RSS': QLineEdit(),
            'CF_RSS': QLineEdit(),
            'sede_RSS': QLineEdit(),
            'firma_RSS': QLineEdit(),
            'img_documento_RSS': QLineEdit(),
            'dichiarazioni_comunicazione_incarichi_RSS': QLineEdit(),
            'dichiarazioni_partecipazione_associazioni_organizzazioni_RSS': QLineEdit(),
            'mail_contatto_RSS': QLineEdit(),
            
            # Protocolli e riferimenti
            'ulteriori_riferimenti_normativi_attuativi_operativi': QLineEdit(),
            'url_gara': QLineEdit(),
            'protocollo_RDA': QLineEdit(),
            'protocollo_richiesta_url': QLineEdit(),
            'protocollo_nomina_RUP': QLineEdit(),
            'protocollo_conflittoint_richiedente': QLineEdit(),
            'protocollo_conflittoint_RUP': QLineEdit(),
            'protocollo_conflittoint_direttore': QLineEdit(),
            'protocollo_allegato2_CIG': QLineEdit(),
            'protocollo_istruttoria_RUP': QLineEdit(),
            'protocollo_DAC': QLineEdit(),
            'protocollo_ordine': QLineEdit()
        }
        
        # Organizza i campi in gruppi logici
        groups = [
            ("Informazioni Generali", [
                'data_oggi', 'data_rda', 'data_scadenza','data_scadenza_offerta',
                'descrizione_servizio_fornitura',
                'breve_descrizione_caratteristiche_prestazioni_acquisizione_bene_servizio',
                'breve_descrizione_motivazione_acquisizione_bene_servizio'
            ]),
            
            ("Clausole e Dichiarazioni", [
                'clausola_cam', 'clausola_servizi_fornitura',
                'dichiarazione_deroga_MEPA', 'dichiarazione_mancata_consip_informatica',
                'dichiarazione_valore_affidamento', 'dichiarazione_motivo_deroga_principio_rotazione'
            ]),
            
            ("Valori Economici", [
                'importo_massimo', 'quantita',
                'importo_oneri_sicurezza', 'importo_oneri_personale'
            ]),
            
            ("Codici e Numeri", [
                'acronimo_progetto', 'numero_CUP', 'numero_CIG',
                'numero_COAN', 'voce_piano_dei_conti', 'voce_costo_COAN',
                'codice_CPV', 'codice_ateco_OE', 'codice_ateco_OE_sec',
                'codice_ateco_OE_dich', 'codice_CNEL', 'estratti_CNEL'
            ]),
            
            ("Informazioni Operatore Economico", [
                'piattaforma_scelta', 'bando_MEPA','riferimento_PAD',
                'dichiarazione_individuazione_OE', 'indirizzo_OE_scelta',
                'legale_rap_OE_scelta', 'sede_OE_scelta', 'piva_OE_scelta'
            ]),
            
            ("Richiedente", [
                'nome_cognome_richiedente', 'data_nascita_richiedente',
                'luogo_nascita_richiedente', 'CF_richiedente', 'sede_richiedente',
                'dichiarazioni_comunicazione_incarichi_richiedente',
                'dichiarazioni_partecipazione_associazioni_organizzazioni_richiedente',
                'mail_contatto_richiedente', 'qualifica_richiedente'
            ]),
            
            ("RUP", [
                'nome_cognome_RUP', 'data_nascita_RUP',
                'luogo_nascita_RUP', 'CF_RUP', 'sede_RUP',
                'dichiarazioni_comunicazione_incarichi_RUP',
                'dichiarazioni_partecipazione_associazioni_organizzazioni_RUP',
                'mail_contatto_RUP'
            ]),
            ("supportoRUP", [
                'nome_cognome_supportoRUP', 'data_nascita_supportoRUP',
                'luogo_nascita_supportoRUP', 'CF_supportoRUP', 'sede_supportoRUP',
                'dichiarazioni_comunicazione_incarichi_supportoRUP',
                'dichiarazioni_partecipazione_associazioni_organizzazioni_supportoRUP',
                'mail_contatto_supportoRUP'
            ]),
            ("Direttore", [
                'nome_cognome_direttore', 'data_nascita_direttore',
                'luogo_nascita_direttore', 'CF_direttore', 'sede_direttore',
                'dichiarazioni_comunicazione_incarichi_direttore',
                'dichiarazioni_partecipazione_associazioni_organizzazioni_direttore',
                'mail_contatto_direttore'
            ]),
            
            ("RSS", [
                'nome_cognome_RSS', 'data_nascita_RSS',
                'luogo_nascita_RSS', 'CF_RSS', 'sede_RSS',
                'dichiarazioni_comunicazione_incarichi_RSS',
                'dichiarazioni_partecipazione_associazioni_organizzazioni_RSS',
                'mail_contatto_RSS'
            ]),
            
            ("Protocolli e Riferimenti", [
                'ulteriori_riferimenti_normativi_attuativi_operativi',
                'url_gara', 'protocollo_RDA', 'protocollo_richiesta_url',
                'protocollo_nomina_RUP', 'protocollo_conflittoint_richiedente',
                'protocollo_conflittoint_RUP', 'protocollo_conflittoint_direttore',
                'protocollo_allegato2_CIG', 'protocollo_istruttoria_RUP',
                'protocollo_DAC', 'protocollo_ordine'
            ]),
            
            ("Generazione Offerte - Campi Principali", [
                'servizio_fornitura', 
                'acronimo_progetto',
                'numero_CUP',
                'oggetto_fornitura_servizio',
                'oggetto_esteso_fornitura_servizio',                                
                'nome_cognome_richiedente', 
                'mail_contatto_richiedente'
            ])
        ]
        
        # Aggiungi i gruppi al layout
        for group_name, fields in groups:
            group_box = QGroupBox(group_name)
            group_layout = QFormLayout()
            
            for field_name in fields:
                label = field_name.replace('_', ' ').title() + ":"
                self.doc_fields[field_name].setPlaceholderText(f"Inserisci {label.lower()}")
                group_layout.addRow(label, self.doc_fields[field_name])
            
            group_box.setLayout(group_layout)
            container_layout.addWidget(group_box)
        
        # Set the container as the scroll area's widget
        scroll_area.setWidget(container)
        
        # Add the scroll area to the main tab layout
        doc_layout.addWidget(scroll_area)
        
        self.tabs.addTab(doc_tab, "Genera Documenti")

    def browse_template_file(self):
        """Open a file dialog to select multiple Word templates"""
        file_names, _ = QFileDialog.getOpenFileNames(
            self, 
            "Seleziona template Word", 
            self.cartella_template,
            "Word Documents (*.docx)"
        )
        if file_names:
            self.template_path.setText("; ".join(file_names))
            self.cartella_template = os.path.dirname(file_names[0])  # Aggiorna la cartella con l'ultima usata
    def browse_output_dir(self):
        dir_name = QFileDialog.getExistingDirectory(
            self, "Seleziona cartella di output"
        )
        if dir_name:
            self.output_dir.setText(dir_name)

    def generate_document(self, source_sheet):
        """Generate Word documents from multiple templates using data from specified sheet"""
        # Get current date in the desired format
        current_date = datetime.now().strftime('%d/%m/%Y')
        
        template_paths = [path.strip() for path in self.template_path.text().split(";") if path.strip()]
        output_dir = self.output_dir.text()
        
        if not template_paths:
            QMessageBox.warning(self, "Attenzione", "Seleziona almeno un template Word.")
            return
        
        if not output_dir:
            output_dir = "A_preventivo"
            self.output_dir.setText(output_dir)
        
        if source_sheet not in self.sheet_data or not self.sheet_data[source_sheet]:
            QMessageBox.warning(self, "Attenzione", f"Nessun dato disponibile dal foglio {source_sheet}")
            return
        
        try:
            # Create output directory if not exists
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # Prepare base context (dati_generali + form fields)
            context = {
                'data_oggi': current_date,
                'data_corrente': current_date
            }
            
            # Add all form fields to context
            for field_name, widget in self.doc_fields.items():
                context[field_name] = widget.text()
            
            # Add data from dati_generali_procedura sheet
            for i in range(self.dati_generali_list.count()):
                item = self.dati_generali_list.item(i)
                item_data = item.data(Qt.ItemDataRole.UserRole)
                context[item_data['name']] = item_data['value']
            
            # Special processing for generazioni_offerte sheet
            if source_sheet == 'generazioni_offerte':
                # Genera un documento per ogni riga selezionata
                selected_items = self.generazioni_offerte_list.selectedItems()
                if not selected_items:
                    QMessageBox.warning(self, "Attenzione", "Seleziona almeno una riga dal foglio generazioni_offerte.")
                    return
                
                generated_files = []
                for item in selected_items:
                    item_data = item.data(Qt.ItemDataRole.UserRole)
                    row_context = context.copy()
                    row_context.update(item_data['data'])
                    
                    # Generate documents for each template
                    for template_path in template_paths:
                        try:
                            # Generate filename with current date
                            cognome = row_context['nome_cognome_richiedente'].split()[-1] if row_context['nome_cognome_richiedente'] else 'documento'
                            progetto = row_context['acronimo_progetto']
                            template_name = os.path.splitext(os.path.basename(template_path))[0]
                            output_filename = f"Richiesta_Offerta_{cognome}_{progetto}_OE_{item_data['row_idx']}_{current_date.replace('/', '-')}.docx"
                            output_path = os.path.join(output_dir, output_filename)
                            
                            # Render and save document
                            doc = DocxTemplate(template_path)
                            doc.render(row_context)
                            doc.save(output_path)
                            generated_files.append(output_path)
                            
                        except Exception as e:
                            QMessageBox.warning(
                                self, 
                                "Attenzione", 
                                f"Errore durante la generazione del documento {template_path}:\n{str(e)}"
                            )
                            continue
                
                if generated_files:
                    success_message = f"Documenti generati con successo il {current_date}:\n\n" + "\n".join(generated_files)
                    QMessageBox.information(self, "Successo", success_message)
                else:
                    QMessageBox.warning(self, "Attenzione", "Nessun documento è stato generato.")
            
            else:
                # Original processing for dati_generali_procedura
                generated_files = []
                for template_path in template_paths:
                    try:
                        # Generate filename with current date
                        cognome = context['nome_cognome_richiedente'].split()[-1] if context['nome_cognome_richiedente'] else 'documento'
                        progetto = context['acronimo_progetto']
                        template_name = os.path.splitext(os.path.basename(template_path))[0]
                        output_filename = f"{template_name}_{cognome}_{progetto}_{current_date.replace('/', '-')}.docx"
                        output_path = os.path.join(output_dir, output_filename)
                        
                        # Render and save document
                        doc = DocxTemplate(template_path)
                        doc.render(context)
                        doc.save(output_path)
                        generated_files.append(output_path)
                        
                    except Exception as e:
                        QMessageBox.warning(
                            self, 
                            "Attenzione", 
                            f"Errore durante la generazione del documento {template_path}:\n{str(e)}"
                        )
                        continue
                
                if generated_files:
                    success_message = f"Documenti generati con successo il {current_date}:\n\n" + "\n".join(generated_files)
                    QMessageBox.information(self, "Successo", success_message)
                else:
                    QMessageBox.warning(self, "Attenzione", "Nessun documento è stato generato.")
                    
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Errore", 
                f"Errore durante la generazione dei documenti il {current_date}:\n{str(e)}"
            )

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = ExcelReaderWindow()
    window.show()
    sys.exit(app.exec())
